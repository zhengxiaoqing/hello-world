import openpyxl
#读取工作表 为wb
wb = openpyxl.load_workbook(r'D:\clouddesk.xlsx')
#创建新表  用create_sheet方法（title是关键字参数）   是为了获得---未登录的人员表
newSheet = wb.create_sheet(title = '过滤后表')
#获得登录sheet    wb的sheetnames的属性是获得所有工作簿名字
sheet1 = wb.get_sheet_by_name(wb.sheetnames[0])
#创建空表，用来保存名字
existNameList = []
#获得登录周所有名字列表
for name in sheet1['A']:
	existNameList.append(name.value)
#获得第二张工作簿---所有员工的表
sheet2 = wb.get_sheet_by_name(wb.sheetnames[1])
#定义新表行数
newRowNum = 2
for rowNum in range(2,len(sheet2['B'])):
#找到没在----已登录表（人数少）中的人员
	if sheet2.cell(row=rowNum,column=2).value not in existNameList:
	#复制前5列到新表
		for colNum in range(1,6):
			newSheet.cell(row=newRowNum, column=colNum).value = sheet2.cell(row=rowNum, column=colNum).value
		#新表行计数器+1
		newRowNum += 1
#保存excel
wb.save(r'D:\clouddesk.xlsx')