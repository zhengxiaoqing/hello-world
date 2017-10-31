import glob
import openpyxl
#将所有人的周报xlsx放入固定文件夹，xls格式不可用
#获得所有周报的一个名字列表
listOfReport = glob.glob(r'D:\documents\testop\*.xlsx')
#打开汇总excel
wbAll = openpyxl.load_workbook(r'D:\documents\huizong.xlsx')
#当前汇总表的活动工作簿---sheetAll
sheetAll = wbAll.get_active_sheet()

#获得所有人员名字的列表
NameList =[]
for cellName in sheetAll['E6': 'E18']:
    for rowOfCellObj in cellName:
        NameList.append(rowOfCellObj.value)

#迭代打开文件夹的每个xlsx	
for reportname in listOfReport:
    print(reportname)
    wbPerson = openpyxl.load_workbook(reportname)
    sheetPerson = wbPerson.get_active_sheet()
	
	#获得每个xlsx的个人名字（名字在D6单元格）
    singleName = sheetPerson['D6'].value
    for rowNum in range(6,19):
		#sheetAll汇总表的第一个名字在E6（第6行，第5列）
		#对比singleName和汇总表中名字
        everyName = sheetAll.cell(row = rowNum, column = 5).value
        if singleName == everyName:
			#汇总表第6列到第9列的单元格4个
            for colNum in range(5,9):
				#从第6行开始到第19行
                sheetAll.cell(row = rowNum , column = colNum +1).value = sheetPerson.cell(row = 6, column = colNum).value
wbAll.save('D:\\documents\\huizong.xlsx')

    




