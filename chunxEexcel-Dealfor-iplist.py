import openpyxl
import os
#将操作文件放到D盘下，名字为worktest.xlsx
os.chdir('D:\\')
i=1    #汇总行计数器
alist=[]
j=1    #用来判断不为空的下一行递增计数器
wb = openpyxl.load_workbook(r'D:\worktest.xlsx')  #打开excel对象
sheetnew = wb.create_sheet(title='new sheet')     #新建一个工作簿
sheet1 = wb.get_sheet_by_name(wb.sheetnames[0])   #获得第一个工作簿
for cell in list(sheet1.columns)[2]:              #获得第三列的所有cell对象
    if cell.value == None:                        #这步很重要，不然复杂度会提升，直接跳过当前行是空的
        continue
    elif cell.value != None:
        # 将vip列ip拆开放入新表第一列（汇总第一列即拆分的单元格）
        sheetnew.cell(row=i, column=1).value = cell.value
        # 将ip列的ip地址填入list，用来实现汇总IP的表现
        alist.append(sheet1.cell(row=cell.row, column=2).value)
        #如果vip列的下一列为None，则将值放入list，一直碰到vip列有值，以此来拆分单元格
        #以下是这段代码最重要的部分
        while sheet1.cell(row=cell.row+j, column=3).value == None:
            alist.append(sheet1.cell(row=cell.row+j, column=2).value)
            j += 1  #j+1，只要下一个值不为空，则将其加入list
            if (cell.row+j) > sheet1.max_row:  #如果行数超过最大行，则break
                break

        j = 1       #j重新计数
        print(alist)
        #格式化列表IP/IP/IP/IP这样的列表
        sheetnew.cell(row=i, column=2).value = '/'.join(alist)
        alist=[]    #清空alist 为下一个不为空的值做准备
        i += 1      #计数器+1
#保存xlsx
wb.save(r'D:\test111.xlsx')